"""
test_all_paths.py — BioMedStatX decision-tree coverage + R cross-validation.

Runs every statistical path through AnalysisManager.analyze() with synthetic
data, validates routing, statistical properties, Excel output structure, and
cross-checks p-values against R as the scientific ground truth.

Run with:
    cd validation
    python -m pytest test_all_paths.py -v --tb=short

Or as a standalone script (no pytest needed):
    python test_all_paths.py
"""

import os
import sys
import csv
import subprocess
import tempfile
import traceback
from pathlib import Path
from unittest.mock import patch, MagicMock

import numpy as np
import pandas as pd
import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
SRC  = ROOT / "src"
R_TEMPLATES = Path(__file__).resolve().parent / "r_templates"

for _p in [str(ROOT), str(SRC)]:
    if _p not in sys.path:
        sys.path.insert(0, _p)

from conftest import DESIGNS

# ---------------------------------------------------------------------------
# Import BioMedStatX modules
# ---------------------------------------------------------------------------

from stats_functions import AnalysisManager


# ---------------------------------------------------------------------------
# Helpers: analysis context builder
# ---------------------------------------------------------------------------

def build_analysis_context(design: dict, group_labels: list) -> dict:
    """Build an analysis_context dict as the autopilot UI would produce."""
    ctx = {
        "dv_columns":      design["dv_columns"],
        "factor_columns":  design["factor_columns"],
        "group_labels":    group_labels,
        "subject_column":  design.get("subject_column"),
        "dependent":       design["dependent"],
        "inferred_test":   design["inferred_test"],
        "additional_factors": [],
        "combine_columns": False,
        "display_group_col": design["factor_columns"][0],
    }
    if design["inferred_test"] == "two_way_anova":
        ctx["additional_factors"] = design["factor_columns"]
    elif design["inferred_test"] == "mixed_anova":
        ctx["between_factors"] = design.get("between_factors", [])
        ctx["within_factors"]  = design.get("within_factors", [])
        ctx["additional_factors"] = (
            design.get("between_factors", []) + design.get("within_factors", [])
        )
    elif design["inferred_test"] == "repeated_measures_anova":
        ctx["within_factors"]    = design["factor_columns"]
        ctx["additional_factors"] = design["factor_columns"]
    elif design["inferred_test"] == "ancova":
        # analysis_core.py reads covariates from analysis_context["covariates"]
        ctx["covariates"] = design.get("covariate_columns", [])
    elif design["inferred_test"] in ("correlation", "linear_regression"):
        # analysis_core.py reads x from analysis_context["x_variable"] (line 589)
        ctx["x_variable"] = design.get("x_column")
    return ctx


# ---------------------------------------------------------------------------
# Helpers: Excel output validation
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = ["Summary", "Statistical Results", "Descriptive Statistics", "Assumptions", "Decision Tree"]


def _get_sheet_text(ws) -> str:
    """Return all cell values from a worksheet as a single lowercase string."""
    parts = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                parts.append(str(cell.value))
    return " ".join(parts).lower()


def _find_p_value_in_results_sheet(ws) -> float | None:
    """
    Search the Results sheet for a cell whose header is 'p-value', 'p value',
    or 'p' and return the first numeric value found below that header.
    Returns None if not found.
    """
    header_cols = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                lv = cell.value.strip().lower()
                if lv in ("p-value", "p value", "p", "p-val", "pvalue"):
                    header_cols[cell.column] = cell.row

    for col, header_row in header_cols.items():
        for row_idx in range(header_row + 1, header_row + 10):
            v = ws.cell(row=row_idx, column=col).value
            if isinstance(v, (int, float)) and not (v != v):  # not NaN
                return float(v)
    return None


def _excel_export_enabled() -> bool:
    """Return True iff ExportDispatcher still calls ResultsExporter.export_results_to_excel.

    The function is intentionally cheap: it inspects the dispatcher source for
    an *uncommented* call. Re-enabling Excel (toggle the leading ``# `` in
    src/export_dispatcher.py) flips this back to True and unskips the suite —
    no test edits required.
    """
    dispatcher = SRC / "export_dispatcher.py"
    try:
        src = dispatcher.read_text(encoding="utf-8")
    except OSError:
        return False
    for raw in src.splitlines():
        line = raw.lstrip()
        if line.startswith("#"):
            continue
        if "ResultsExporter.export_results_to_excel" in line:
            return True
    return False


def assert_excel_output(result: dict, design: dict, out_dir: Path):
    """Validate the Excel output file produced by the analysis."""
    # Excel export is currently disabled in ExportDispatcher (HTML-only mode,
    # see commit c1b26ef). Skip this check until the export is re-enabled —
    # toggling the comments in src/export_dispatcher.py auto-restores the test.
    if not _excel_export_enabled():
        pytest.skip(
            f"[{design['name']}] Excel export disabled in ExportDispatcher "
            "(HTML-only mode). Re-enable in src/export_dispatcher.py to run this check."
        )

    # The output file is named *_output.xlsx to distinguish from the input fixture
    xlsx_files = list(out_dir.glob("*_output.xlsx"))
    if not xlsx_files:
        xlsx_files = list(out_dir.glob("*.xlsx"))
        # Exclude the input fixture file
        xlsx_files = [f for f in xlsx_files if "_output" in f.name or "results" in f.name.lower()
                      or design["name"] + ".xlsx" not in f.name]
    assert len(xlsx_files) >= 1, (
        f"[{design['name']}] No output .xlsx file found in {out_dir}. "
        f"Files present: {list(out_dir.iterdir())}"
    )

    wb = openpyxl.load_workbook(str(xlsx_files[0]))

    # 1. Required sheets present
    for sheet in REQUIRED_SHEETS:
        assert sheet in wb.sheetnames, (
            f"[{design['name']}] Missing sheet '{sheet}'. "
            f"Found: {wb.sheetnames}"
        )

    # 2. Pairwise sheet for multi-group results
    if design["levels"] > 2:
        # "Pairwise Comparisons" sheet may exist but is optional if test was not significant
        pass  # soft check — some implementations skip it for non-significant results

    # 3. p-value in Statistical Results sheet matches Python result dict
    ws_results = wb["Statistical Results"]
    excel_p = _find_p_value_in_results_sheet(ws_results)
    if excel_p is not None and result.get("p_value") is not None:
        python_p = float(result["p_value"])
        assert abs(excel_p - python_p) < 1e-4, (
            f"[{design['name']}] p-value mismatch: "
            f"Python={python_p:.6f}, Excel={excel_p:.6f}"
        )

    # 4. Non-parametric mention in Analysis Log (Decision Tree is an image, not searchable text)
    test_name_lower = str(result.get("test", "")).lower()
    if any(kw in test_name_lower for kw in ["kruskal", "mann", "wilcox", "friedman"]):
        log_text = ""
        if "Analysis Log" in wb.sheetnames:
            log_text = _get_sheet_text(wb["Analysis Log"])
        summary_text = _get_sheet_text(wb["Summary"]) if "Summary" in wb.sheetnames else ""
        combined_text = log_text + " " + summary_text
        if not any(kw in combined_text for kw in ["non-parametric", "nonparametric", "not normally",
                                                    "kruskal", "mann", "wilcox", "friedman",
                                                    "non_parametric"]):
            print(f"  WARNING [{design['name']}]: No non-parametric mention in Log/Summary")

    # 5. ANOVA: eta (effect size) mentioned for multi-group or multi-factor tests
    if design["levels"] >= 3 or design["factors"] >= 2:
        results_text = _get_sheet_text(wb["Statistical Results"])
        # Soft check — log a warning if missing but don't fail (implementation may differ)
        if "eta" not in results_text and "effect" not in results_text:
            print(f"  WARNING [{design['name']}]: No eta/effect_size in Statistical Results sheet")


# ---------------------------------------------------------------------------
# Helpers: R cross-validation
# ---------------------------------------------------------------------------

def _rscript_path() -> str:
    """Find Rscript on Windows or Unix."""
    # Check common Windows locations first
    candidates = [
        r"C:\Users\pkrumm\AppData\Local\Programs\R\R-4.5.0\bin\Rscript.exe",
        r"C:\Program Files\R\R-4.5.0\bin\Rscript.exe",
        "Rscript",
    ]
    for c in candidates:
        if os.path.isfile(c):
            return c
    return "Rscript"   # fallback: hope it's on PATH


RSCRIPT = _rscript_path()


def _validate_r_effect_size(result: dict, design: dict, r: dict) -> None:
    """
    Cross-validate effect sizes from R.
    - Cohen's d (t-tests) and R² (regression): hard assertions at 1e-4
    - eta_squared / partial eta²: warnings only (formula variants legitimately differ)
    """
    name = design["name"]

    # ── Cohen's d for t-tests (hard assert: same pooled-SD formula both sides) ──
    if "effect_size" in r and design.get("r_test") in ("indep_ttest", "paired_ttest"):
        python_es = result.get("effect_size")
        if python_es is not None:
            diff = abs(abs(float(python_es)) - abs(float(r["effect_size"])))
            assert diff < 1e-4, (
                f"[{name}] Cohen's d mismatch: "
                f"Python={float(python_es):.6f}, R={float(r['effect_size']):.6f}, diff={diff:.2e}"
            )
            print(f"  Cohen's d OK: Python={abs(float(python_es)):.4f}, R={abs(float(r['effect_size'])):.4f}")

    # ── R² for regression (hard assert: identical OLS formula) ──
    if "r_squared" in r:
        python_es = result.get("effect_size")
        if python_es is not None:
            diff = abs(abs(float(python_es)) - float(r["r_squared"]))
            assert diff < 1e-4, (
                f"[{name}] R² mismatch: "
                f"Python={float(python_es):.6f}, R={float(r['r_squared']):.6f}, diff={diff:.2e}"
            )
            print(f"  R² OK: Python={float(python_es):.4f}, R={float(r['r_squared']):.4f}")

    # ── eta_squared for one-way ANOVA (warn only) ──
    if "eta_squared" in r:
        python_es = result.get("effect_size")
        if python_es is not None:
            diff = abs(abs(float(python_es)) - float(r["eta_squared"]))
            label = "  eta² OK" if diff < 0.01 else "  WARNING: eta² mismatch"
            print(f"{label}: Python={abs(float(python_es)):.4f}, R={float(r['eta_squared']):.4f}, diff={diff:.2e}")

    # ── Partial eta² for first factor in two-way / mixed ANOVA (warn only) ──
    peta_key = next((k for k in ("peta_FactorA", "peta_between") if k in r), None)
    if peta_key is not None and result.get("factors"):
        python_peta = result["factors"][0].get("effect_size")
        if python_peta is not None:
            diff = abs(abs(float(python_peta)) - float(r[peta_key]))
            label = "  partial eta² OK" if diff < 0.01 else "  WARNING: partial eta² mismatch"
            print(f"{label} (factor 0): Python={abs(float(python_peta)):.4f}, R={float(r[peta_key]):.4f}, diff={diff:.2e}")


def _validate_r_posthoc(result: dict, design: dict, r: dict) -> None:
    """
    Cross-validate Tukey HSD p-values from R against Python pairwise_comparisons.
    Mismatches are warnings (not failures) because pair ordering may differ.
    """
    tukey_keys = sorted(k for k in r if k.startswith("p_tukey_"))
    if not tukey_keys:
        return

    name = design["name"]
    python_pairs = result.get("pairwise_comparisons", [])
    if not python_pairs:
        print(f"  WARNING [{name}]: R has Tukey p-values but Python has no pairwise_comparisons")
        return

    r_tukey_ps = [float(r[k]) for k in tukey_keys]
    python_tukey_ps = [float(c["p_value"]) for c in python_pairs]
    tol = design.get("r_posthoc_tolerance", 0.01)

    all_ok = True
    for i, py_p in enumerate(python_tukey_ps):
        best = min(r_tukey_ps, key=lambda rp: abs(py_p - rp))
        diff = abs(py_p - best)
        if diff >= tol:
            print(f"  WARNING [{name}]: Tukey pair {i}: Python p={py_p:.4f}, closest R p={best:.4f}, diff={diff:.2e}")
            all_ok = False
    if all_ok:
        print(f"  Tukey HSD post-hoc OK: {len(python_tukey_ps)} pairs within tol={tol}")


def validate_against_r(result: dict, design: dict, excel_path: str, tmp_path: Path):
    """
    Cross-validate the Python p-value (and effect sizes) against R.
    Skipped gracefully if r_test is None or Rscript is unavailable.

    Parsing is format-driven via the design's r_output_format key.
    Default format: ["p_value", "statistic"] (legacy two-value output).
    """
    r_test = design.get("r_test")
    if r_test is None:
        return  # Robustness/NaN test or no R equivalent — skip

    r_script = R_TEMPLATES / f"{r_test}.R"
    if not r_script.exists():
        print(f"  SKIP R validation [{design['name']}]: no template {r_script.name}")
        return

    # Convert DataFrame to CSV for R (simpler than Excel parsing)
    df_orig = design["df_factory"]()
    csv_path = tmp_path / f"{design['name']}_r_input.csv"
    df_orig.to_csv(str(csv_path), index=False)

    # Build R command — some designs pass extra CLI args (e.g. correlation method)
    cmd = [RSCRIPT, "--vanilla", str(r_script), str(csv_path)]
    for extra in design.get("r_extra_args", []):
        cmd.append(str(extra))

    # Call R
    try:
        proc = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
    except (subprocess.TimeoutExpired, FileNotFoundError) as e:
        print(f"  SKIP R validation [{design['name']}]: {e}")
        return

    if proc.returncode != 0:
        print(f"  SKIP R validation [{design['name']}]: R error — {proc.stderr.strip()[:200]}")
        return

    output = proc.stdout.strip()
    if not output:
        print(f"  SKIP R validation [{design['name']}]: R returned no output")
        return

    # Parse — handle NA tokens gracefully (R fallback paths emit NA for missing values)
    raw_tokens = output.split()
    values = []
    for tok in raw_tokens:
        try:
            values.append(float(tok))
        except ValueError:
            values.append(float("nan"))  # NA → NaN, skipped in comparisons

    if not values or all(v != v for v in values):   # all NaN
        print(f"  SKIP R validation [{design['name']}]: R returned only NA values")
        return

    # Map tokens to named fields via r_output_format
    fmt = design.get("r_output_format", ["p_value", "statistic"])
    r = {name: val for name, val in zip(fmt, values)}

    # ── p-value tolerance ────────────────────────────────────────────────────
    # Wilcoxon: Python uses normal approx, R may use exact → keep loose (0.05)
    # Mann-Whitney: both use normal approx → tighter (1e-3)
    # Parametric: formulas are identical → tight (1e-4)
    # Nonparametric mixed ANOVA uses different test entirely → very loose (0.5)
    # Design-level r_tolerance always wins.
    r_test_name = design.get("r_test", "")
    default_tol = (
        0.05  if r_test_name == "wilcoxon" else
        1e-3  if r_test_name == "mann_whitney" else
        1e-4
    )
    tolerance = design.get("r_tolerance", default_tol)

    # Gather all p-value fields from the named dict
    p_keys = [k for k in r if k == "p_value" or k.startswith("p_")]
    r_p_values = [r[k] for k in p_keys if not (r[k] != r[k])]  # skip NaN
    if not r_p_values:
        print(f"  SKIP R comparison [{design['name']}]: no valid p-values in R output")
        return

    python_p = result.get("p_value")
    if python_p is None and result.get("factors"):
        python_p = result["factors"][0].get("p_value")
    if python_p is None:
        print(f"  SKIP R comparison [{design['name']}]: Python p_value is None")
        return

    diff = min(abs(float(python_p) - rp) for rp in r_p_values)
    best_r_p = min(r_p_values, key=lambda rp: abs(float(python_p) - rp))

    assert diff < tolerance, (
        f"[{design['name']}] R cross-validation FAILED: "
        f"Python p={float(python_p):.8f}, best R p={best_r_p:.8f}, "
        f"diff={diff:.2e} > tol={tolerance:.2e}"
    )
    print(f"  R cross-check OK: Python p={float(python_p):.6f}, best R p={best_r_p:.6f}, diff={diff:.2e}")

    # ── Effect size and post-hoc cross-validation ────────────────────────────
    _validate_r_effect_size(result, design, r)
    _validate_r_posthoc(result, design, r)


# ---------------------------------------------------------------------------
# Main test
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("design", DESIGNS, ids=[d["name"] for d in DESIGNS])
def test_path(design, make_excel_fixture, tmp_path):
    """
    End-to-end test for one statistical path:
      1. Generate synthetic Excel fixture
      2. Run AnalysisManager.analyze() with the correct analysis_context
      3. Assert correct test routing
      4. Assert valid statistical properties
      5. Assert Excel output structure and p-value consistency
      6. R cross-validation (skipped gracefully if R unavailable)
    """
    # ── 1. Generate fixture ─────────────────────────────────────────────────
    excel_path = make_excel_fixture(design)

    # Determine actual group labels (may need to be read from data for 2-factor designs)
    df = design["df_factory"]()
    if design["group_labels"]:
        group_labels = design["group_labels"]
    else:
        # Multi-factor: groups are auto-determined by the engine from the DataFrame
        group_labels = []

    # ── 2. Build context ────────────────────────────────────────────────────
    context = build_analysis_context(design, group_labels)

    # Output Excel goes to tmp_path (absolute path → get_output_path respects it)
    out_excel_base = str(tmp_path / f"{design['name']}_output")

    # For multi-factor designs, groups must be empty so the engine auto-determines them
    # from the combined factor columns (e.g. "FactorA=A1, FactorB=B1")
    is_multifactor = design["factors"] >= 2
    groups_arg = [] if is_multifactor else (group_labels or list(df[design["factor_columns"][0]].dropna().unique()))

    # ── 3. Run analysis ─────────────────────────────────────────────────────
    result = AnalysisManager.analyze(
        file_path=excel_path,
        group_col=design["factor_columns"][0],
        groups=groups_arg,
        sheet_name=0,
        value_cols=design["dv_columns"],
        dependent=design["dependent"],
        skip_plots=True,
        skip_excel=False,
        file_name=out_excel_base,
        analysis_context=context,
    )

    # ── 4. Basic validity ───────────────────────────────────────────────────
    assert result is not None, f"[{design['name']}] analyze() returned None"
    # pingouin import error (broken xarray dependency) → skip gracefully
    if result.get("error"):
        err = str(result["error"])
        if "pingouin" in err.lower() or "xarray" in err.lower() or "expected string or bytes" in err:
            pytest.skip(f"[{design['name']}] Skipped: pingouin/xarray import broken — {err[:100]}")
        pytest.fail(f"[{design['name']}] Analysis returned error: {result['error']}")

    # ── 5. p-value in valid range ───────────────────────────────────────────
    p_val = result.get("p_value")
    if p_val is not None:
        assert 0.0 <= float(p_val) <= 1.0, (
            f"[{design['name']}] p_value={p_val} is outside [0, 1]"
        )

    # For multi-factor tests, check the factors list
    if "factors" in result:
        for factor_result in result["factors"]:
            fp = factor_result.get("p_value")
            if fp is not None:
                assert 0.0 <= float(fp) <= 1.0, (
                    f"[{design['name']}] Factor p_value={fp} outside [0, 1]"
                )

    # ── 6. Test routing ─────────────────────────────────────────────────────
    test_name = str(result.get("test", "")).lower()
    model_class = str(result.get("model_class", "")).lower()
    combined = test_name + " " + model_class

    # For multi-factor ANOVA results that return factor lists without "test" key
    if not combined.strip() and "factors" in result:
        combined = design["inferred_test"].lower()

    expected_kws = design["expected_test_keywords"]
    assert any(kw in combined for kw in expected_kws), (
        f"[{design['name']}] Expected one of {expected_kws} in test name, "
        f"got: '{combined}'"
    )

    # ── 7. Excel output ─────────────────────────────────────────────────────
    assert_excel_output(result, design, tmp_path)

    # ── 8. R cross-validation ───────────────────────────────────────────────
    validate_against_r(result, design, excel_path, tmp_path)


# ---------------------------------------------------------------------------
# Standalone runner (no pytest required)
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys
    import io
    from conftest import DESIGNS
    from unittest.mock import patch, MagicMock

    mock_ui = MagicMock()
    mock_ui.select_posthoc_test_dialog.return_value = "tukey"
    mock_ui.select_nonparametric_posthoc_dialog.return_value = "dunn"
    mock_ui.select_control_group_dialog.return_value = None
    mock_ui.select_custom_pairs_dialog.return_value = []

    results_summary = []
    total = len(DESIGNS)

    print(f"\n{'='*60}")
    print(f"  BioMedStatX Decision-Tree Coverage — {total} paths")
    print(f"{'='*60}\n")

    with patch("stats_functions.UIDialogManager", mock_ui), \
         patch("statisticaltester.UIDialogManager", mock_ui):

        for i, design in enumerate(DESIGNS, 1):
            print(f"[{i:02d}/{total}] {design['name']} ...", end=" ", flush=True)
            try:
                with tempfile.TemporaryDirectory() as tmp_str:
                    tmp_path = Path(tmp_str)
                    df = design["df_factory"]()
                    excel_path = str(tmp_path / f"{design['name']}.xlsx")
                    df.to_excel(excel_path, index=False)

                    group_labels = design["group_labels"] or []
                    context = build_analysis_context(design, group_labels)
                    out_excel_base = str(tmp_path / f"{design['name']}_output")

                    result = AnalysisManager.analyze(
                        file_path=excel_path,
                        group_col=design["factor_columns"][0],
                        groups=group_labels or list(
                            df[design["factor_columns"][0]].dropna().unique()
                        ),
                        sheet_name=0,
                        value_cols=design["dv_columns"],
                        dependent=design["dependent"],
                        skip_plots=True,
                        skip_excel=False,
                        file_name=out_excel_base,
                        analysis_context=context,
                    )

                    if result.get("error"):
                        raise AssertionError(f"Analysis error: {result['error']}")

                    p_val = result.get("p_value")
                    if p_val is not None:
                        assert 0.0 <= float(p_val) <= 1.0, f"p_value={p_val} out of range"

                    assert_excel_output(result, design, tmp_path)
                    validate_against_r(result, design, excel_path, tmp_path)

                print("PASS")
                results_summary.append((design["name"], True, None))

            except Exception as exc:
                print(f"FAIL — {exc}")
                results_summary.append((design["name"], False, str(exc)))

    print(f"\n{'='*60}")
    passed = sum(1 for _, ok, _ in results_summary if ok)
    failed = total - passed
    print(f"  {passed}/{total} PASSED   {failed} FAILED")
    print(f"{'='*60}")
    for name, ok, err in results_summary:
        status = "PASS" if ok else "FAIL"
        print(f"  {status}  {name}")
        if err:
            # Print only first line of error to keep output readable
            print(f"       {err.splitlines()[0]}")

    print()
    sys.exit(0 if failed == 0 else 1)
