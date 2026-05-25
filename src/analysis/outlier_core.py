import os
from datetime import datetime

import numpy as np
import pandas as pd
from scipy.stats import t
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.lazy_imports import get_seaborn

OUTLIER_IMPORTS_AVAILABLE = True
try:
    import matplotlib.pyplot as plt
except ImportError:
    OUTLIER_IMPORTS_AVAILABLE = False

class OutlierDetector:
    """
    Detect outliers in grouped data using Grubbs' Test or Dixon's Q-Test.
    Loads an Excel table with columns ['Group', 'Value'], 
    converts all values (German decimal numbers with comma) to float,
    performs Grubbs or Dixon tests iteratively or once for each group separately
    and marks found outliers.
    """

    def __init__(self, df, group_col, value_col):
        """
        Initializes the OutlierDetector with an already loaded DataFrame.
        
        Parameters:
        -----------
        df : pandas.DataFrame
            DataFrame with the data
        group_col : str
            Name of the group column
        value_col : str
            Name of the values column
        """
        print(f"DEBUG: Initializing OutlierDetector with columns: group_col={group_col}, value_col={value_col}")
        print(f"DEBUG: DataFrame shape: {df.shape}, columns: {df.columns.tolist()}")
        
        if not OUTLIER_IMPORTS_AVAILABLE:
            raise ImportError("Required packages for outlier detection are not available. "
                              "Please install: pip install outliers pingouin openpyxl")
        
        # Create copy of DataFrame
        self.df = df.copy()
        self.group_col = group_col
        self.value_col = value_col
        self.active_test = None  # Track which test was run
        self.debug_log = []  # Initialize debug log

        # Ensure columns are present
        if group_col not in self.df.columns or value_col not in self.df.columns:
            raise ValueError(f"Columns '{group_col}' and '{value_col}' must be present in the DataFrame.")

        # Add initialization info to log
        self.debug_log.append("*** OUTLIER DETECTION INITIALIZATION ***")
        self.debug_log.append(f"DataFrame shape: {df.shape}")
        self.debug_log.append(f"Group column: {group_col}")
        self.debug_log.append(f"Value column: {value_col}")
        self.debug_log.append(f"Available columns: {df.columns.tolist()}")

        # Convert values column to float (if necessary)
        self._convert_values_to_float()
        
        # Show group statistics after initialization
        self.debug_log.append("\n=== GROUP STATISTICS ===")
        self.debug_log.append(f"Number of groups in data: {self.df[group_col].nunique()}")
        
        for group_name, group_data in self.df.groupby(group_col):
            group_stats = self._calculate_group_statistics(group_name, group_data)
            self.debug_log.extend(group_stats)

    def _convert_values_to_float(self):
        """
        Converts the values column to float.
        Handles German decimal separators (comma instead of period).
        """
        try:
            self.debug_log.append("\n=== VALUE CONVERSION ===")
            self.debug_log.append(f"Converting column '{self.value_col}' to float")
            
            # Check if the column exists
            if self.value_col not in self.df.columns:
                error_msg = f"Column '{self.value_col}' not found in DataFrame. Available columns: {list(self.df.columns)}"
                self.debug_log.append(f"ERROR: {error_msg}")
                raise ValueError(error_msg)
            
            self.debug_log.append(f"Sample values before conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type before conversion: {self.df[self.value_col].dtype}")
            
            # Check if column is already numeric
            if pd.api.types.is_numeric_dtype(self.df[self.value_col]):
                self.debug_log.append("Column is already numeric, no conversion needed")
                return
            
            # Convert German decimal numbers
            self.df[self.value_col] = (
                self.df[self.value_col]
                    .astype(str)
                    .str.replace('.', '', regex=False)   # Remove thousand separators
                    .str.replace(',', '.', regex=False)  # Comma → Period
                    .astype(float)
            )
            
            self.debug_log.append(f"Sample values after conversion: {self.df[self.value_col].head(3).tolist()}")
            self.debug_log.append(f"Data type after conversion: {self.df[self.value_col].dtype}")
            self.debug_log.append(f"Value range after conversion: {self.df[self.value_col].min():.4f} to {self.df[self.value_col].max():.4f}")
        
        except Exception as e:
            error_message = f"Conversion failed: {str(e)}"
            self.debug_log.append(f"ERROR: {error_message}")
            raise ValueError(f"Error converting values column '{self.value_col}' to float: {str(e)}")
        
    def _calculate_group_statistics(self, group_name, group_data):
        """Calculate basic statistics for a group and return as log entries."""
        stats = []
        values = group_data[self.value_col].dropna()
        stats.append(f"\n--- Group '{group_name}' Statistics ---")
        stats.append(f"  Count: {len(values)}")
        
        if len(values) > 0:
            stats.append(f"  Min: {values.min():.4f}")
            stats.append(f"  Max: {values.max():.4f}")
            stats.append(f"  Mean: {values.mean():.4f}")
            stats.append(f"  Median: {values.median():.4f}")
            stats.append(f"  Std Dev: {values.std():.4f}")
        else:
            stats.append("  No valid values in this group")
        
        return stats
    
    @staticmethod
    def _grubbs_iterative(vals: np.ndarray,
                           alpha: float = 0.05,
                           two_sided: bool = True):
        """
        Iteratively remove the most extreme point (if G_obs > G_crit),
        repeat until no further outliers are detected.
        Returns list of relative indices into the original vals array.
        """
        vals = np.asarray(vals, dtype=float)
        # track the original positions
        idxs = np.arange(vals.size)
        out_rel_idxs = []
        
        while vals.size >= 3:
            G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                vals, alpha=alpha, two_sided=two_sided
            )
            if G_obs <= G_crit:
                break
            # record and remove
            out_rel_idxs.append(int(idxs[rel_idx]))
            mask = np.ones(vals.size, dtype=bool)
            mask[rel_idx] = False
            vals = vals[mask]
            idxs = idxs[mask]
        
        return out_rel_idxs

    @staticmethod
    def _grubbs_statistic(x: np.ndarray,
                          alpha: float = 0.05,
                          two_sided: bool = True):
        """
        Compute Grubbs' G statistic and critical value for array x.
        Returns (G_obs, G_crit, outlier_rel_index).
        """
        n = x.size
        if n < 3:
            raise ValueError("Grubbs' test requires at least 3 values")
        
        mu = x.mean()
        s  = x.std(ddof=1)
        # maximum absolute deviation
        diffs = np.abs(x - mu)
        rel_idx = int(np.argmax(diffs))
        G_obs = diffs[rel_idx] / s
        
        # two‐sided splits alpha into 2 tails
        tail = 2 if two_sided else 1
        # student-t critical value
        t_crit = t.ppf(1 - alpha/(tail * n), df=n-2)
        # critical G
        G_crit = ((n - 1) / np.sqrt(n)
                  * np.sqrt(t_crit**2 / ( (n - 2) + t_crit**2 )))
        
        return G_obs, G_crit, rel_idx

    def run_grubbs(self, alpha: float = 0.05, iterate: bool = True):
        self.debug_log.append("\n=== GRUBBS OUTLIER DETECTION ===")
        self.debug_log.append(f"alpha={alpha}, iterate={iterate}")

        self.df["Grubbs_Outlier"] = False
        self.active_test = "Grubbs"
        total_out = 0

        for gname, gdf in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Group '{gname}' ({len(gdf)}) ---")

            # original indices and clean values
            idxs = gdf.index.values
            vals = gdf[self.value_col].astype(float).values
            valid = ~np.isnan(vals)
            idxs_valid = idxs[valid]
            vals_valid = vals[valid]

            if len(vals_valid) < 3:
                self.debug_log.append("  n < 3 → skipping Grubbs")
                continue

            # choose iterative vs. single removal
            if iterate:
                rel_outs = OutlierDetector._grubbs_iterative(
                    vals_valid, alpha=alpha, two_sided=True
                )
            else:
                G_obs, G_crit, rel_idx = OutlierDetector._grubbs_statistic(
                    vals_valid, alpha=alpha, two_sided=True
                )
                rel_outs = [rel_idx] if G_obs > G_crit else []

            if not rel_outs:
                self.debug_log.append("  No outliers detected")
            else:
                # map back to global indices
                global_outs = idxs_valid[rel_outs].tolist()
                for gi in global_outs:
                    val = self.df.at[gi, self.value_col]
                    self.debug_log.append(f"  Outlier at idx={gi}, value={val:.4f}")
                self.df.loc[global_outs, "Grubbs_Outlier"] = True
                total_out += len(global_outs)
                self.debug_log.append(f"  ==> {len(global_outs)} outliers in group")

        self.debug_log.append("\n=== GRUBBS SUMMARY ===")
        self.debug_log.append(f"Total outliers: {total_out}")

    def run_grubbs_test(self, alpha: float = 0.05, iterate: bool = True):
        """
        Alias for run_grubbs method for compatibility.
        
        Parameters:
        -----------
        alpha : float
            Significance level for the test (default: 0.05)
        iterate : bool
            Whether to iteratively remove outliers (default: True)
        """
        return self.run_grubbs(alpha=alpha, iterate=iterate)

    def run_mod_z_score(self, threshold=3.5, iterate=False):
        """
        Performs Modified Z-Score outlier detection for each group.
        
        Parameters:
        -----------
        threshold : float
            Threshold for modified Z-scores to be considered outliers (default: 3.5)
        iterate : bool
            Whether to iteratively remove outliers and recompute scores
        """
        self.debug_log.append("\n=== MODIFIED Z-SCORE OUTLIER DETECTION EXECUTION ===")
        self.debug_log.append(f"Test parameters: threshold={threshold}, iterate={iterate}")
        self.debug_log.append("Test principle: Modified Z-Score uses median absolute deviation (MAD) for robustness against outliers")
        
        # Create the Modified Z-Score column and mark this test as active
        self.df['ModZ_Outlier'] = False
        self.active_test = 'ModZ'
        
        total_outliers = 0
        
        for group_name, group_df in self.df.groupby(self.group_col):
            self.debug_log.append(f"\n--- Processing group '{group_name}' ---")
            self.debug_log.append(f"Group size: {len(group_df)} observations")
            
            indices = group_df.index.tolist()
            values = group_df[self.value_col].values.copy()
            
            # Filter out NaN values
            valid_mask = ~np.isnan(values)
            valid_values = values[valid_mask]
            valid_indices = [idx for mask, idx in zip(valid_mask, indices) if mask]
            
            if len(valid_values) == 0:
                self.debug_log.append(f"Group '{group_name}' has no valid values (all NaN), skipping")
                continue
                
            outlier_indices = []
            
            # Log basic group statistics
            mean_val = np.mean(valid_values)
            median_val = np.median(valid_values)
            self.debug_log.append(f"Group statistics: min={np.min(valid_values):.3f}, max={np.max(valid_values):.3f}, mean={mean_val:.3f}, median={median_val:.3f}")
            
            def detect_single_round(vals, indices_list):
                if len(vals) <= 1:
                    return [], "No outliers found (too few values for MAD calculation)"
                    
                # Calculate median and MAD
                median = np.median(vals)
                mad = np.median(np.abs(vals - median))
                
                # Avoid division by zero
                if mad == 0:
                    return [], "MAD = 0, cannot compute modified Z-scores"
                    
                # Calculate modified Z-scores
                mod_z_scores = 0.6745 * (vals - median) / mad
                
                # Find outliers
                outlier_mask = np.abs(mod_z_scores) > threshold
                outlier_idx = [idx for mask, idx in zip(outlier_mask, indices_list) if mask]
                
                # Log this round
                self.debug_log.append("  Modified Z-Score Analysis:")
                self.debug_log.append(f"    Median: {median:.4f}")
                self.debug_log.append(f"    MAD: {mad:.4f}")
                self.debug_log.append(f"    Threshold: {threshold} (absolute value)")
                
                if len(outlier_idx) > 0:
                    self.debug_log.append(f"    {len(outlier_idx)} outliers found in this round")
                    return outlier_idx, None
                else:
                    self.debug_log.append("    No outliers found in this round")
                    return [], "No outliers found (all modified Z-scores within threshold)"
            
            # Iterative outlier detection
            vals_copy = valid_values.copy()
            idx_copy = valid_indices.copy() 
            round_count = 0
            
            while True:
                round_count += 1
                self.debug_log.append(f"  Testing round {round_count} with {len(vals_copy)} values")
                
                round_outliers, stop_message = detect_single_round(vals_copy, idx_copy)
                
                if round_outliers:
                    outlier_indices.extend(round_outliers)
                    if iterate:
                        # Remove outliers for next iteration
                        keep_mask = ~np.isin(idx_copy, round_outliers)
                        vals_copy = vals_copy[keep_mask]
                        idx_copy = [idx for keep, idx in zip(keep_mask, idx_copy) if keep]
                    else:
                        self.debug_log.append("  No more iterations requested, terminating test")
                        break
                else:
                    self.debug_log.append(f"  {stop_message if stop_message else 'No more outliers found'}, terminating test")
                    break
                    
                if len(vals_copy) <= 1:
                    self.debug_log.append("  Too few values left for MAD calculation, terminating test")
                    break
            
            # Mark all found indices in main DataFrame
            if outlier_indices:
                self.df.loc[outlier_indices, 'ModZ_Outlier'] = True
                total_outliers += len(outlier_indices)
                self.debug_log.append(f"  RESULT: {len(outlier_indices)} outliers found in group '{group_name}'")
            else:
                self.debug_log.append(f"  RESULT: No outliers found in group '{group_name}'")
        
        # Final summary
        self.debug_log.append("\n=== MODIFIED Z-SCORE TEST SUMMARY ===")
        self.debug_log.append(f"Total outliers detected across all groups: {total_outliers}")
        self.debug_log.append("Test completed successfully")


    def save_results(self, output_path):
        """
        Outputs the result to a new Excel file while preserving formulas.
        Only columns for actually performed tests are displayed.
        """
        self.debug_log.append("\n=== SAVING RESULTS ===")
        self.debug_log.append(f"Output path: {output_path}")
        self.debug_log.append(f"Active test: {self.active_test}")
        
        # Store the original path - it's crucial to use absolute paths
        output_path = os.path.abspath(output_path)
        self.debug_log.append(f"Absolute output path: {output_path}")
        
        # Store original directory but don't change it
        original_cwd = os.getcwd()
        self.debug_log.append(f"Current working directory: {original_cwd}")
        
        try:
            if self.active_test == 'Grubbs':
                outlier_count = self.df['Grubbs_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Grubbs method: {outlier_count}")
            elif self.active_test == 'ModZ':
                outlier_count = self.df['ModZ_Outlier'].sum()
                self.debug_log.append(f"Total outliers found with Modified Z-Score method: {outlier_count}")
            
            # Check if file exists and contains formulas we need to preserve
            file_exists = os.path.exists(output_path)
            
            if file_exists:
                # Create a backup of the original file with a timestamp
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_dir = os.path.dirname(output_path)
                base_name = os.path.splitext(os.path.basename(output_path))[0]
                backup_path = os.path.join(output_dir, f"{base_name}_{timestamp}_backup.xlsx")
                import shutil
                shutil.copy2(output_path, backup_path)
                self.debug_log.append(f"Created backup of existing file at: {backup_path}")
                
                # Create a new file with a different name for our results
                results_path = os.path.join(output_dir, f"{base_name}_outliers.xlsx")
                self.debug_log.append(f"Creating new results file at: {results_path}")
            else:
                # Just use the original path if the file doesn't exist
                results_path = output_path
            
            # Create new workbook for our results
            wb = Workbook()

            # 1) Sheet "Raw_Data_with_Outlier_Marking"
            ws_raw = wb.active
            ws_raw.title = "Raw_Data_with_Outlier_Marking"

            # Include only columns for the test that was run
            test_columns = []
            if self.active_test == 'Grubbs' and 'Grubbs_Outlier' in self.df.columns:
                test_columns.append('Grubbs_Outlier')
            elif self.active_test == 'ModZ' and 'ModZ_Outlier' in self.df.columns:
                test_columns.append('ModZ_Outlier')

            # Write headers
            headers = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws_raw.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Write data with outlier highlighting
            outlier_count_by_group = {}
            for row_idx, (_, row) in enumerate(self.df.iterrows(), start=2):
                group = row[self.group_col]
                if group not in outlier_count_by_group:
                    outlier_count_by_group[group] = 0
                    
                # Column 1: Group
                group_value = str(group) if group is not None else ""
                ws_raw.cell(row=row_idx, column=1, value=group_value)
                
                # Column 2: Value - mark red if outlier
                value = row[self.value_col]
                if pd.isna(value) or not isinstance(value, (int, float)):
                    value = 0.0
                
                cell_value = ws_raw.cell(row=row_idx, column=2, value=float(value))
                
                # Check for outlier and mark with red highlighting
                is_outlier = any(row.get(col, False) for col in test_columns)
                if is_outlier:
                    outlier_count_by_group[group] += 1
                    # Red font and light red background
                    cell_value.font = Font(color="FF0000", bold=True)
                    cell_value.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    self.debug_log.append(f"Marking outlier: Group={group}, Value={value:.4f}")

            # 2) Sheet "Cleaned"
            ws_clean = wb.create_sheet(title="Cleaned")
            headers_clean = [self.group_col, self.value_col]
            for col_idx, header in enumerate(headers_clean, start=1):
                cell = ws_clean.cell(row=1, column=col_idx, value=str(header))
                cell.font = Font(bold=True)

            # Filter: no outliers from active test
            outlier_mask = pd.Series(False, index=self.df.index)
            for test_col in test_columns:
                outlier_mask |= self.df[test_col]
            
            cleaned_df = self.df[~outlier_mask]
            self.debug_log.append(f"Cleaned data has {len(cleaned_df)} rows (removed {len(self.df) - len(cleaned_df)} outliers)")
            
            for row_idx, (_, row) in enumerate(cleaned_df.iterrows(), start=2):
                group_value = str(row[self.group_col]) if row[self.group_col] is not None else ""
                value = float(row[self.value_col]) if not pd.isna(row[self.value_col]) else 0.0
                
                ws_clean.cell(row=row_idx, column=1, value=group_value)
                ws_clean.cell(row=row_idx, column=2, value=value)

            # 3) Summary sheet with statistics
            ws_summary = wb.create_sheet(title="Summary")
            ws_summary.cell(row=1, column=1, value="Outlier Detection Summary").font = Font(bold=True, size=14)
            
            # Test parameters
            ws_summary.cell(row=3, column=1, value="Test parameters:").font = Font(bold=True)
            ws_summary.cell(row=4, column=1, value="Test type:")
            ws_summary.cell(row=4, column=2, value=str(self.active_test))
            
            # Add test-specific parameters
            if self.active_test == 'Grubbs':
                ws_summary.cell(row=5, column=1, value="Grubbs alpha level:")
                ws_summary.cell(row=5, column=2, value="0.05")  # Default value, adjust if needed
            elif self.active_test == 'ModZ':
                ws_summary.cell(row=5, column=1, value="Z-score threshold:")
                ws_summary.cell(row=5, column=2, value="3.5")  # Default value, adjust if needed
            
            # Group statistics
            row_pos = 7
            ws_summary.cell(row=row_pos, column=1, value="Group statistics:").font = Font(bold=True)
            row_pos += 1
            
            # Headers for statistics table
            stat_headers = ["Group", "Count", "Min", "Max", "Mean", "Std", "Outliers"]
            for col_idx, header in enumerate(stat_headers, start=1):
                cell = ws_summary.cell(row=row_pos, column=col_idx, value=header)
                cell.font = Font(bold=True)
                
            row_pos += 1
            for group_name, group_df in self.df.groupby(self.group_col):
                values = group_df[self.value_col].dropna()
                
                ws_summary.cell(row=row_pos, column=1, value=str(group_name))
                ws_summary.cell(row=row_pos, column=2, value=len(group_df))
                ws_summary.cell(row=row_pos, column=3, value=float(values.min()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=4, value=float(values.max()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=5, value=float(values.mean()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=6, value=float(values.std()) if len(values) > 0 else 0.0)
                ws_summary.cell(row=row_pos, column=7, value=outlier_count_by_group.get(group_name, 0))
                row_pos += 1

            # 4) Enhanced Debug Log sheet with better formatting
            ws_log = wb.create_sheet(title="Debug_Log")
            header_cell = ws_log.cell(row=1, column=1)
            header_cell.value = "Outlier Detection Debug Log"
            header_cell.font = Font(bold=True, size=16, color="0066CC")

            subheader_cell = ws_log.cell(row=2, column=1)
            subheader_cell.value = "This sheet contains detailed information about the outlier detection process."
            subheader_cell.font = Font(italic=True)

            # Set column width for better readability
            ws_log.column_dimensions['A'].width = 150

            # Write debug log entries with enhanced formatting and formula protection
            for idx, log_entry in enumerate(self.debug_log, start=4):
                log_text = str(log_entry)[:32000]
                cell = ws_log.cell(row=idx, column=1)
                cell.value = log_text
                
                # Force inlineStr data type for any text starting with "="
                if log_text.startswith("="):
                    cell.data_type = 'inlineStr'
                
                # Enhanced formatting based on content
                if log_text.startswith("==="):
                    # Main section headers - blue background, white text
                    cell.font = Font(bold=True, size=14, color="FFFFFF")
                    cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
                elif log_text.startswith("---"):
                    # Group headers - light blue background, dark text
                    cell.font = Font(bold=True, size=12, color="000080")
                    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                elif "ERROR:" in log_text:
                    # Errors - red background, white text
                    cell.font = Font(color="FFFFFF", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
                elif "OUTLIER DETECTED" in log_text:
                    # Outlier detection - yellow background, red text
                    cell.font = Font(color="CC0000", bold=True, size=11)
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                elif "RESULT:" in log_text:
                    # Results - green background, dark text
                    cell.font = Font(color="006600", bold=True, size=11)
                    cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
                elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                    # Test round information - light gray background
                    cell.font = Font(size=10, italic=True)
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                elif log_text.startswith("    "):
                    # Detailed information - smaller font, indented
                    cell.font = Font(size=9, color="666666")
                elif log_text.startswith("  "):
                    # General information - smaller font
                    cell.font = Font(size=10)

            # Add visualization
            self._add_single_visualization_sheet(wb, self)
            
            # Save the new workbook - Make sure to use the absolute path
            wb.save(results_path)
            self.debug_log.append(f"Results saved to: {results_path}")
            
            # If we created a separate file, inform user about both files
            if file_exists:
                self.debug_log.append(f"Original file with formulas preserved at: {output_path}")
                self.debug_log.append(f"Backup copy created at: {backup_path}")
                print(f"NOTE: To preserve formulas, results were saved to a new file: {results_path}")
                print(f"      Original file with formulas intact: {output_path}")
        
        except Exception as e:
            error_msg = f"Error saving results: {str(e)}"
            self.debug_log.append(f"ERROR: {error_msg}")
            print(f"ERROR: {error_msg}")
            raise

    def get_summary(self):
        """
        Creates a summary of found outliers.
        """
        self.debug_log.append("\n=== CREATING SUMMARY ===")
        
        total_rows = len(self.df)
        grubbs_outliers = self.df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in self.df.columns else 0
        modz_outliers = self.df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in self.df.columns else 0
        
        self.debug_log.append(f"Total rows: {total_rows}")
        self.debug_log.append(f"Grubbs outliers: {grubbs_outliers}")
        self.debug_log.append(f"Modified Z-Score outliers: {modz_outliers}")
        
        any_outliers = 0
        
        # Check which columns exist before combining
        if 'Grubbs_Outlier' in self.df.columns and 'ModZ_Outlier' in self.df.columns:
            any_outliers = (self.df['Grubbs_Outlier'] | self.df['ModZ_Outlier']).sum()
        elif 'Grubbs_Outlier' in self.df.columns:
            any_outliers = grubbs_outliers
        elif 'ModZ_Outlier' in self.df.columns:
            any_outliers = modz_outliers
        
        self.debug_log.append(f"Combined outliers: {any_outliers}")
        
        # Outliers per group
        group_summary = {}
        for group_name, group_df in self.df.groupby(self.group_col):
            group_total = len(group_df)
            group_grubbs = group_df['Grubbs_Outlier'].sum() if 'Grubbs_Outlier' in group_df.columns else 0
            group_modz = group_df['ModZ_Outlier'].sum() if 'ModZ_Outlier' in group_df.columns else 0
            
            group_any = 0
            if 'Grubbs_Outlier' in group_df.columns and 'ModZ_Outlier' in group_df.columns:
                group_any = (group_df['Grubbs_Outlier'] | group_df['ModZ_Outlier']).sum()
            elif 'Grubbs_Outlier' in group_df.columns:
                group_any = group_grubbs
            elif 'ModZ_Outlier' in group_df.columns:
                group_any = group_modz
            
            self.debug_log.append(f"Group '{group_name}': total={group_total}, outliers={group_any}")
            
            group_summary[group_name] = {
                'total': group_total,
                'grubbs_outliers': group_grubbs,
                'modz_outliers': group_modz,
                'any_outliers': group_any
            }
        
        summary = {
            'total_rows': total_rows,
            'grubbs_outliers': grubbs_outliers,
            'modz_outliers': modz_outliers,
            'any_outliers': any_outliers,
            'groups': group_summary
        }
        
        self.debug_log.append(f"Summary completed: {summary}")
        return summary

    @staticmethod
    def run_multi_dataset_outlier_detection(df, group_col, dataset_columns, alpha=0.05, 
                                            iterate=False, run_grubbs=True, run_modz=True, 
                                            grubbs_alpha=0.05, modz_threshold=3.5,
                                            output_path="multi_dataset_outlier_results.xlsx"):
        """
        Run outlier detection on multiple datasets (columns) and create a combined Excel output.
        """
        print(f"DEBUG: Current working directory before export: {os.getcwd()}")
        print("DEBUG: Starting multi-dataset outlier detection")
        print(f"DEBUG: Datasets to analyze: {dataset_columns}")
        print(f"DEBUG: Group column: {group_col}")
        
        all_results = {}
        failed_datasets = {}
        temp_files = []  # Track temporary files for cleanup
        
        # Create main workbook for combined results
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)
        
        try:
            # Analyze each dataset
            for i, dataset_col in enumerate(dataset_columns):
                print(f"DEBUG: Analyzing dataset {i+1}/{len(dataset_columns)}: {dataset_col}")
                
                try:
                    # Create detector for this dataset
                    detector = OutlierDetector(
                        df=df.copy(),
                        group_col=group_col,
                        value_col=dataset_col
                    )
                    
                    # Run requested tests
                    if run_grubbs:
                        print(f"DEBUG: Running Grubbs test for {dataset_col}")
                        detector.run_grubbs(alpha=grubbs_alpha, iterate=iterate)
                    
                    if run_modz:
                        print(f"DEBUG: Running Modified Z-Score test for {dataset_col}")
                        detector.run_mod_z_score(threshold=modz_threshold, iterate=iterate)
                    
                    # Store results
                    all_results[dataset_col] = {
                        'detector': detector,
                        'summary': detector.get_summary()
                    }
                    
                    # Add sheets to combined workbook
                    OutlierDetector._add_dataset_to_workbook(wb, detector, dataset_col, run_grubbs, run_modz)
                    print(f"DEBUG: Successfully analyzed {dataset_col}")
                    
                except Exception as e:
                    error_msg = f"Error analyzing {dataset_col}: {str(e)}"
                    failed_datasets[dataset_col] = error_msg
                    print(f"DEBUG: ERROR: {error_msg}")
                        
            # Create summary sheet
            OutlierDetector._create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns)
            
            # Add visualization sheet with swarm plots
            visual_temp_files = OutlierDetector._add_visualization_sheet(wb, all_results, dataset_columns)
            if visual_temp_files:
                temp_files.extend(visual_temp_files)
            
            # Save combined workbook
            wb.save(output_path)
            print(f"DEBUG: Combined results saved to: {output_path}")
            
            # Return summary
            return {
                "type": "multi_dataset_outlier_detection",
                "successful_datasets": list(all_results.keys()),
                "failed_datasets": failed_datasets,
                "total_datasets": len(dataset_columns),
                "output_file": output_path,
                "summary": {
                    "success_count": len(all_results),
                    "failure_count": len(failed_datasets),
                    "success_rate": f"{len(all_results)/len(dataset_columns)*100:.1f}%"
                }
            }
            
        except Exception as e:
            error_msg = f"Critical error in multi-dataset analysis: {str(e)}"
            print(f"DEBUG: CRITICAL ERROR: {error_msg}")
            raise RuntimeError(error_msg)
            
        finally:
            # Clean up temporary files
            for temp_file in temp_files:
                try:
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                        print(f"DEBUG: Cleaned up temporary file: {temp_file}")
                except Exception as e:
                    print(f"DEBUG: Error cleaning up file {temp_file}: {str(e)}")
    
    @staticmethod
    def _add_dataset_to_workbook(wb, detector, dataset_name, run_grubbs, run_modz):
        """Add analysis sheet for a single dataset with debug log and summary."""
        
        # Create sheet name (Excel sheet names are limited to 31 characters)
        safe_name = dataset_name.replace(' ', '_')[:25]  # Leave room for suffixes
        
        # Create single analysis sheet
        ws = wb.create_sheet(title=f"{safe_name}_Analysis")
        
        # Title
        ws.cell(row=1, column=1, value=f"Outlier Analysis: {dataset_name}").font = Font(bold=True, size=16, color="0066CC")
        
        # Summary section
        summary = detector.get_summary()
        row = 3
        
        ws.cell(row=row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
        row += 1
        
        ws.cell(row=row, column=1, value="Dataset:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=dataset_name)
        row += 1
        
        ws.cell(row=row, column=1, value="Total data points:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['total_rows'])
        row += 1
        
        if run_grubbs:
            ws.cell(row=row, column=1, value="Grubbs outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['grubbs_outliers'])
            row += 1
            
        if run_modz:
            ws.cell(row=row, column=1, value="ModZ outliers:").font = Font(bold=True)
            ws.cell(row=row, column=2, value=summary['modz_outliers'])
            row += 1
        
        ws.cell(row=row, column=1, value="Total outliers:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=summary['any_outliers'])
        row += 2
        
        # Group statistics
        ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
        row += 1
        
        group_headers = ["Group", "Total", "Outliers", "Clean"]
        for col_idx, header in enumerate(group_headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row += 1
        
        for group_name, group_stats in summary['groups'].items():
            ws.cell(row=row, column=1, value=str(group_name))
            ws.cell(row=row, column=2, value=group_stats['total'])
            ws.cell(row=row, column=3, value=group_stats['any_outliers'])
            ws.cell(row=row, column=4, value=group_stats['total'] - group_stats['any_outliers'])
            row += 1
        
        row += 2
        
        # Debug Log section
        ws.cell(row=row, column=1, value="DEBUG LOG").font = Font(bold=True, size=14, color="0066CC")
        row += 1
        
        ws.cell(row=row, column=1, value="Detailed analysis log:").font = Font(italic=True)
        row += 1
        
        # Set column width for debug log
        ws.column_dimensions['A'].width = 40
        
        # Write debug log entries with enhanced formatting
        for log_entry in detector.debug_log:
            log_text = str(log_entry)[:32000]
            cell = ws.cell(row=row, column=1)
            cell.value = log_text
            
            # Force inlineStr data type for any text starting with "="
            if log_text.startswith("="):
                cell.data_type = 'inlineStr'
            
            # Enhanced formatting based on content
            if log_text.startswith("==="):
                # Main section headers - blue background, white text
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            elif log_text.startswith("---"):
                # Group headers - light blue background, dark text
                cell.font = Font(bold=True, size=10, color="000080")
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            elif "ERROR:" in log_text:
                # Errors - red background, white text
                cell.font = Font(color="FFFFFF", bold=True, size=9)
                cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type="solid")
            elif "OUTLIER DETECTED" in log_text:
                # Outlier detection - yellow background, red text
                cell.font = Font(color="CC0000", bold=True, size=9)
                cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
            elif "RESULT:" in log_text:
                # Results - green background, dark text
                cell.font = Font(color="006600", bold=True, size=9)
                cell.fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
            elif log_text.startswith("  Testing round") or log_text.startswith("  Running"):
                # Test round information - light gray background
                cell.font = Font(size=8, italic=True)
                cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            elif log_text.startswith("    "):
                # Detailed information - smaller font, indented
                cell.font = Font(size=8, color="666666")
            elif log_text.startswith("  "):
                # General information - smaller font
                cell.font = Font(size=9)
            
            row += 1

        # Set other column widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    @staticmethod
    def _create_multi_summary_sheet(wb, all_results, failed_datasets, dataset_columns):
        """Create a comprehensive summary sheet for all datasets."""
        ws = wb.create_sheet(title="Data_without_outlier", index=0)  # Insert as first sheet
        
        # Get the original DataFrame from the first successful result
        original_df = None
        group_col = None
        for dataset_col, result_data in all_results.items():
            if 'detector' in result_data:
                original_df = result_data['detector'].df
                group_col = result_data['detector'].group_col
                break
        
        if original_df is None:
            # Fallback if no data available
            ws.cell(row=1, column=1, value="No data available").font = Font(bold=True, color="FF0000")
            return
        
        if group_col is None:
            ws.cell(row=1, column=1, value="Group column not found").font = Font(bold=True, color="FF0000")
            return
        
        # Create headers: Sample + all dataset columns
        headers = [group_col] + dataset_columns
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        
        # Get all unique combinations of group values that exist in the original data
        # We need to preserve the original row structure
        row_idx = 2
        
        # Iterate through the original DataFrame to maintain the exact structure
        for orig_idx, orig_row in original_df.iterrows():
            group_value = orig_row[group_col]
            
            # Column 1: Group/Sample name
            ws.cell(row=row_idx, column=1, value=str(group_value))
            
            # For each dataset column, get the value and check if it's an outlier
            for col_idx, dataset_col in enumerate(dataset_columns, start=2):
                if dataset_col in all_results:
                    detector = all_results[dataset_col]['detector']
                    
                    # Get the value for this specific row
                    if orig_idx < len(detector.df):
                        row_data = detector.df.iloc[orig_idx]
                        value = row_data[dataset_col]
                        
                        # Check if this specific row is marked as an outlier
                        is_grubbs_outlier = row_data.get('Grubbs_Outlier', False) if 'Grubbs_Outlier' in detector.df.columns else False
                        is_modz_outlier = row_data.get('ModZ_Outlier', False) if 'ModZ_Outlier' in detector.df.columns else False
                        is_outlier = is_grubbs_outlier or is_modz_outlier
                        
                        # Add the value to the cell
                        if pd.isna(value):
                            cell = ws.cell(row=row_idx, column=col_idx, value="NaN")
                        else:
                            cell = ws.cell(row=row_idx, column=col_idx, value=float(value))
                            
                            # Mark outliers with red highlighting
                            if is_outlier:
                                cell.font = Font(color="FF0000", bold=True)
                                cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    else:
                        # Row doesn't exist in this detector's data
                        ws.cell(row=row_idx, column=col_idx, value="N/A")
                else:
                    # Dataset failed or not available
                    ws.cell(row=row_idx, column=col_idx, value="N/A")
            
            row_idx += 1
        
        # Set column widths
        for col in range(1, len(headers) + 1):
            if col <= 26:  # A-Z
                ws.column_dimensions[chr(64 + col)].width = 15
            else:  # AA, AB, etc.
                ws.column_dimensions[f"A{chr(64 + col - 26)}"].width = 15
                
    @staticmethod
    def _add_visualization_sheet(wb, all_results, dataset_columns):
        """Add visualization sheet with swarm plots showing outliers."""
        from openpyxl.drawing.image import Image
        sns = get_seaborn()
        import tempfile
        # import os  # Already imported at top
        
        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")
        
        # Set title
        ws.cell(row=1, column=1, value="Outlier Visualization").font = Font(bold=True, size=16, color="0066CC")
        ws.cell(row=2, column=1, value="Visual representation of data with outliers highlighted").font = Font(italic=True)
        
        # Track row position and temp files
        row = 4
        temp_files = []
        
        # Track if we generated any plots successfully
        plots_added = False
        
        # For each dataset, create a swarm plot with modern styling
        for dataset_col in dataset_columns:
            if dataset_col not in all_results:
                continue
            
            # Get detector for this dataset
            result_data = all_results[dataset_col]
            if 'detector' not in result_data:
                continue
            
            detector = result_data['detector']
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col
            
            # Determine which outlier column to use
            outlier_col = None
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            
            if not outlier_col:
                continue
                
            try:
                # Add dataset name as header
                title_text = f"Dataset: {dataset_col}"
                title_cell = ws.cell(row=row, column=1)
                title_cell.value = title_text
                title_cell.font = Font(bold=True, size=14)
                if title_text.startswith("="):
                    title_cell.data_type = 'inlineStr'
                row += 1
                
                # Add test method used
                ws.cell(row=row, column=1, value=f"Method: {test_name}")
                row += 1
                
                # Get outlier count
                outlier_count = df[outlier_col].sum()
                total_count = len(df)
                ws.cell(row=row, column=1, value=f"Found {outlier_count} outliers out of {total_count} data points ({outlier_count/total_count:.1%})")
                row += 2
                
                # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
                # Modern, minimalist style
                sns.set_style("white")  
                plt.rcParams.update({
                    "axes.edgecolor": "#333333",
                    "axes.linewidth": 1.0,
                    "xtick.color": "#333333",
                    "ytick.color": "#333333",
                    "font.size": 12,
                    "axes.titlesize": 14,
                    "axes.labelsize": 12,
                })

                # Larger figure size with less whitespace
                fig, ax = plt.subplots(figsize=(10, 7))
                fig.patch.set_facecolor("white")

                # Boxplot in subtle light gray (no outlier symbols)
                sns.boxplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    fliersize=0,
                    width=0.5,
                    palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                    ax=ax
                )

                # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
                # Depending on dataset size: stripplot with jitter
                if len(df) < 100:
                    sns.swarmplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                        size=10,
                        edgecolor="white",
                        linewidth=0.8,
                        ax=ax
                    )
                else:
                    sns.stripplot(
                        x=group_col,
                        y=value_col,
                        data=df,
                        hue=outlier_col,
                        palette={False: "#555555", True: "#007AFF"},
                        size=8,
                        jitter=0.25,
                        alpha=0.8,
                        dodge=False,
                        ax=ax
                    )

                # Remove top and right spines ("despine")
                sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

                # Axis labels
                ax.set_xlabel(group_col, fontsize=12, color="#333333")
                ax.set_ylabel(value_col, fontsize=12, color="#333333")
                ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

                # Legend only for outliers, no frame, modern positioning
                handles, labels = ax.get_legend_handles_labels()
                # labels come in order [False, True] -> filter accordingly
                new_handles, new_labels = [], []
                for h, lab in zip(handles, labels):
                    if lab == "False":
                        new_handles.append(h)
                        new_labels.append("Normal")
                    elif lab == "True":
                        new_handles.append(h)
                        new_labels.append("Outlier")
                ax.legend(
                    new_handles,
                    new_labels,
                    title="Status",
                    loc="upper right",
                    frameon=False,
                    fontsize=12,
                    title_fontsize=13
                )

                # Subtle grid on y-axis (almost invisible)
                ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
                ax.set_axisbelow(True)

                # Tight layout (less margin)
                plt.tight_layout(pad=1)
                
                # Create a temporary file path with high DPI for crisp rendering
                fd, temp_path = tempfile.mkstemp(suffix='.png')
                os.close(fd)
                temp_files.append(temp_path)
                
                # Save with higher DPI for better quality
                plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
                plt.close(fig)
                
                # Add the image to the worksheet
                img = Image(temp_path)
                # Set image size (in pixels)
                img.width = 1000
                img.height = 700
                ws.add_image(img, f'A{row}')
                
                # Mark that we added at least one plot
                plots_added = True
                
                # Move row position down to accommodate the image
                row += 32
                row += 2  # Add some space between plots
                
            except Exception as e:
                error_cell = ws.cell(row=row, column=1)
                error_cell.value = f"Error creating visualization for {dataset_col}: {str(e)}"
                error_cell.data_type = 'inlineStr'  # Ensure error messages aren't interpreted as formulas
                row += 3
        
        # If no plots were added, add a message
        if not plots_added:
            ws.cell(row=row, column=1, value="No visualizations could be generated.")
            row += 1
        
        # Set column width for better display
        ws.column_dimensions['A'].width = 120
        
        # Return list of temp files for later cleanup
        return temp_files
    
    @staticmethod
    def _add_single_visualization_sheet(wb, detector, dataset_name=None):
        """Add visualization sheet with swarm plot showing outliers in modern, minimalist style."""
        from openpyxl.drawing.image import Image
        sns = get_seaborn()
        import tempfile
        # import os  # Already imported at top

        # Create a visualization sheet
        ws = wb.create_sheet(title="Visualization")

        # Set title (Excel sheet)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = "Outlier Visualization"
        title_cell.font = Font(bold=True, size=16, color="0066CC")

        subtitle_cell = ws.cell(row=2, column=1)
        subtitle_cell.value = "Visual representation with highlighted outliers"
        subtitle_cell.font = Font(italic=True)

        # Start row for the image
        row = 4
        temp_path = None

        try:
            df = detector.df
            group_col = detector.group_col
            value_col = detector.value_col

            # Which outlier column?
            if 'Grubbs_Outlier' in df.columns:
                outlier_col = 'Grubbs_Outlier'
                test_name = "Grubbs Test"
            elif 'ModZ_Outlier' in df.columns:
                outlier_col = 'ModZ_Outlier'
                test_name = "Modified Z-Score Test"
            else:
                # If no results
                no_results_cell = ws.cell(row=row, column=1)
                no_results_cell.value = "No outlier results available."
                no_results_cell.data_type = 'inlineStr'
                return None

            # Subheader on the Excel sheet
            title_text = f"Dataset: {dataset_name or value_col}"
            title2_cell = ws.cell(row=row, column=1)
            title2_cell.value = title_text
            title2_cell.font = Font(bold=True, size=14)
            if title_text.startswith("="):
                title2_cell.data_type = 'inlineStr'
            row += 1

            method_cell = ws.cell(row=row, column=1)
            method_cell.value = f"Outlier Method: {test_name}"
            row += 1

            # Outlier counter
            outlier_count = int(df[outlier_col].sum())
            total_count = len(df)
            count_cell = ws.cell(row=row, column=1)
            count_cell.value = f"Found outliers: {outlier_count} of {total_count} ({outlier_count/total_count:.1%})"
            row += 2

            # ─── Matplotlib/Seaborn ‐ Plot ────────────────────────────────────────────
            # Modern, minimalist style
            sns.set_style("white")
            plt.rcParams.update({
                "axes.edgecolor": "#333333",
                "axes.linewidth": 1.0,
                "xtick.color": "#333333",
                "ytick.color": "#333333",
                "font.size": 12,
                "axes.titlesize": 14,
                "axes.labelsize": 12,
            })

            # Larger figure size with less whitespace
            fig, ax = plt.subplots(figsize=(10, 7))
            fig.patch.set_facecolor("white")

            # Boxplot in subtle light gray (no outlier symbols)
            sns.boxplot(
                x=group_col,
                y=value_col,
                data=df,
                fliersize=0,
                width=0.5,
                palette=["#DDDDDD"] * len(df[group_col].unique()),   # Light gray box
                ax=ax
            )

            # Scatter/Swarm plot: normal points in dark gray, outliers in bright blue
            # Depending on dataset size: stripplot with jitter
            if len(df) < 100:
                sns.swarmplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},  # Bright blue for outliers
                    size=10,
                    edgecolor="white",
                    linewidth=0.8,
                    ax=ax
                )
            else:
                sns.stripplot(
                    x=group_col,
                    y=value_col,
                    data=df,
                    hue=outlier_col,
                    palette={False: "#555555", True: "#007AFF"},
                    size=8,
                    jitter=0.25,
                    alpha=0.8,
                    dodge=False,
                    ax=ax
                )

            # Remove top and right spines ("despine")
            sns.despine(ax=ax, top=True, right=True, left=False, bottom=False)

            # Axis labels
            ax.set_xlabel(group_col, fontsize=12, color="#333333")
            ax.set_ylabel(value_col, fontsize=12, color="#333333")
            ax.set_title(f"{value_col} by Group (Outliers Highlighted)", fontsize=14, color="#333333", pad=15)

            # Legend only for outliers, no frame, modern positioning
            handles, labels = ax.get_legend_handles_labels()
            # labels come in order [False, True] -> filter accordingly
            new_handles, new_labels = [], []
            for h, lab in zip(handles, labels):
                if lab == "False":
                    new_handles.append(h)
                    new_labels.append("Normal")
                elif lab == "True":
                    new_handles.append(h)
                    new_labels.append("Outlier")
            ax.legend(
                new_handles,
                new_labels,
                title="Status",
                loc="upper right",
                frameon=False,
                fontsize=12,
                title_fontsize=13
            )

            # Subtle grid on y-axis (almost invisible)
            ax.grid(axis="y", color="#f0f0f0", linestyle="-", linewidth=0.7)
            ax.set_axisbelow(True)

            # Tight layout (less margin)
            plt.tight_layout(pad=1)

            # Save as PNG temporarily
            fd, temp_path = tempfile.mkstemp(suffix=".png")
            os.close(fd)
            plt.savefig(temp_path, dpi=200, bbox_inches="tight", facecolor="white")
            plt.close(fig)

            # Insert image into Excel sheet
            img = Image(temp_path)
            img.width = 1000  # in pixels
            img.height = 700
            ws.add_image(img, f"A{row}")

            # ─── Group Statistics Table ───────────────────────────────────────────
            row += 32  # Leave space for the image
            ws.cell(row=row, column=1, value="Group Statistics:").font = Font(bold=True)
            row += 1

            # Table headers
            stats_headers = ["Group", "Count", "Mean", "StdDev", "Median", "Min", "Max", "Outliers"]
            for col_idx, header in enumerate(stats_headers, start=1):
                cell = ws.cell(row=row, column=col_idx, value=header)
                cell.font = Font(bold=True)
            row += 1

            # Values per group
            for group_name, group_df in df.groupby(group_col):
                vals = group_df[value_col].dropna()
                out_count = int(group_df[outlier_col].sum())

                # Group
                grp_cell = ws.cell(row=row, column=1)
                grp_val = str(group_name)
                grp_cell.value = grp_val
                if grp_val.startswith("="):
                    grp_cell.data_type = 'inlineStr'

                # Numeric cells
                ws.cell(row=row, column=2, value=len(vals))
                ws.cell(row=row, column=3, value=float(vals.mean()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=4, value=float(vals.std()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=5, value=float(vals.median()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=6, value=float(vals.min()) if len(vals) > 0 else 0.0)
                ws.cell(row=row, column=7, value=float(vals.max()) if len(vals) > 0 else 0.0)

                out_cell = ws.cell(row=row, column=8, value=out_count)
                # Highlight if there are outliers
                if out_count > 0:
                    out_cell.font = Font(color="FF0000", bold=True)
                    for c in range(1, 9):
                        ws.cell(row=row, column=c).fill = PatternFill(
                            start_color="FFEEEE",
                            end_color="FFEEEE",
                            fill_type="solid"
                        )
                row += 1

        except Exception as e:
            # If an error occurs during plotting
            err_cell = ws.cell(row=row, column=1)
            err_cell.value = f"Error in visualization: {str(e)}"
            err_cell.data_type = 'inlineStr'

        # Adjust column widths
        for col in range(1, 9):
            if col == 1:
                ws.column_dimensions[chr(64 + col)].width = 30
            else:
                ws.column_dimensions[chr(64 + col)].width = 15

        return temp_path
